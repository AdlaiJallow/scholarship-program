"""
Reporting and Excel/CSV export (system specification §19, §20). Export
runs synchronously for the MVP scaffold's default request/response cycle;
production should move large exports behind the Celery worker already
wired up for email (apps.notifications.tasks) once real data volume makes
a request-time .xlsx build slow enough to matter.
"""

import csv
from datetime import date, timedelta

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.permissions import HasRolePermission
from apps.audit.models import AuditLog
from apps.audit.services import log_action

from .models import Application
from .views import ApplicationFilter


class ReportsSummaryView(APIView):
    permission_classes = [HasRolePermission]
    required_permission = "reports.view"

    def get(self, request):
        from apps.accounts.models import Student

        today = timezone.now().date()
        month_start = today.replace(day=1)
        qs = Application.objects.exclude(status=Application.Status.NOT_STARTED)

        counts = {
            "total_scholarship_holders": Student.objects.count(),
            "applications_received": qs.count(),
            "pending_verification": qs.filter(status=Application.Status.IN_PROGRESS).count(),
            "under_review": qs.filter(status=Application.Status.UNDER_REVIEW).count(),
            "approved": qs.filter(status=Application.Status.APPROVED).count(),
            "rejected": qs.filter(status=Application.Status.REJECTED).count(),
            "awaiting_student_response": qs.filter(
                status__in=[Application.Status.ADDITIONAL_INFO_REQUIRED, Application.Status.RESUBMISSION_REQUIRED]
            ).count(),
            "submitted_today": qs.filter(submitted_at__date=today).count(),
            "submitted_this_month": qs.filter(submitted_at__date__gte=month_start).count(),
        }
        return Response(counts)


class ReportsBreakdownView(APIView):
    permission_classes = [HasRolePermission]
    required_permission = "reports.view"

    DIMENSIONS = {
        "country": "scholarship__country__name",
        "institution": "scholarship__institution__name",
        "scholarship_type": "scholarship__scholarship_type__name",
        "academic_level": "scholarship__program__academic_level",
        "rejection_reason": "rejection_reason",
    }

    def get(self, request):
        from django.db.models import Count

        dimension = request.query_params.get("by", "country")
        field = self.DIMENSIONS.get(dimension)
        if field is None:
            return Response(
                {"detail": f"Unsupported breakdown dimension. Choose one of: {', '.join(self.DIMENSIONS)}."},
                status=400,
            )
        qs = (
            Application.objects.exclude(status=Application.Status.NOT_STARTED)
            .values(field)
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        return Response([{"label": row[field] or "Unspecified", "count": row["count"]} for row in qs])


class ReportsProcessingTimeView(APIView):
    permission_classes = [HasRolePermission]
    required_permission = "reports.view"

    def get(self, request):
        decided = Application.objects.filter(submitted_at__isnull=False, decided_at__isnull=False)
        durations = [(a.decided_at - a.submitted_at) for a in decided]
        if not durations:
            return Response({"average_processing_time_hours": None, "decided_count": 0})
        average = sum(durations, timedelta()) / len(durations)
        total = decided.count()
        completed = Application.objects.filter(status=Application.Status.APPROVED).count() + Application.objects.filter(
            status=Application.Status.REJECTED
        ).count()
        received = Application.objects.exclude(status=Application.Status.NOT_STARTED).count()
        completion_rate = round(completed / received, 4) if received else None
        return Response(
            {
                "average_processing_time_hours": round(average.total_seconds() / 3600, 1),
                "decided_count": total,
                "verification_completion_rate": completion_rate,
            }
        )


EXPORT_COLUMNS = [
    ("Application ID", lambda a: str(a.id)),
    ("Scholarship ID", lambda a: a.scholarship.scholarship_reference_id),
    ("Student Name", lambda a: a.scholarship.student.full_name),
    ("Institution", lambda a: a.scholarship.institution.name),
    ("Country", lambda a: a.scholarship.country.name),
    ("Program", lambda a: a.scholarship.program.name if a.scholarship.program else ""),
    ("Scholarship Type", lambda a: a.scholarship.scholarship_type.name),
    ("Submission Date", lambda a: a.submitted_at.date().isoformat() if a.submitted_at else ""),
    ("Verification Status", lambda a: a.get_status_display()),
    ("Approval/Rejection Date", lambda a: a.decided_at.date().isoformat() if a.decided_at else ""),
    ("Officer Responsible", lambda a: a.decided_by.full_name if a.decided_by else ""),
    ("Rejection Reason", lambda a: a.get_rejection_reason_display() if a.rejection_reason else ""),
    ("Last Updated", lambda a: a.updated_at.date().isoformat()),
]
# National ID numbers and raw document files are deliberately excluded here
# per the data-privacy architecture (system specification §17, §20).


class ExportView(APIView):
    permission_classes = [HasRolePermission]
    required_permission = "exports.run"

    def get(self, request):
        # Deliberately "export_format", not "format": DRF's content
        # negotiation already reserves the "format" query parameter for
        # choosing the *response* renderer (?format=json) — reusing it here
        # made DRF 404 on any value it didn't recognize as a renderer.
        file_format = request.query_params.get("export_format", "xlsx")
        queryset = ApplicationFilter(
            request.query_params, queryset=Application.objects.exclude(status=Application.Status.NOT_STARTED)
        ).qs.select_related(
            "scholarship__student",
            "scholarship__institution",
            "scholarship__country",
            "scholarship__program",
            "scholarship__scholarship_type",
            "decided_by",
        )

        log_action(
            AuditLog.Action.EXPORT_GENERATED,
            actor=request.user,
            metadata={"format": file_format, "row_count": queryset.count(), "filters": dict(request.query_params)},
            request=request,
        )

        if file_format == "csv":
            return self._csv_response(queryset)
        return self._xlsx_response(queryset)

    def _csv_response(self, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="verification_export_{date.today()}.csv"'
        writer = csv.writer(response)
        writer.writerow([label for label, _ in EXPORT_COLUMNS])
        for application in queryset:
            writer.writerow([getter(application) for _, getter in EXPORT_COLUMNS])
        return response

    def _xlsx_response(self, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Verification Records"

        header_fill = PatternFill(start_color="0B4F4A", end_color="0B4F4A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_index, (label, _) in enumerate(EXPORT_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_index, value=label)
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"

        for row_index, application in enumerate(queryset, start=2):
            for col_index, (_, getter) in enumerate(EXPORT_COLUMNS, start=1):
                sheet.cell(row=row_index, column=col_index, value=getter(application))

        for col_index, (label, _) in enumerate(EXPORT_COLUMNS, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = max(16, len(label) + 4)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="verification_export_{date.today()}.xlsx"'
        workbook.save(response)
        return response
